from openpyxl import Workbook
from openpyxl.styles import Font

from io import BytesIO
from flask import make_response


def makeResponse(content, fileName):
    resp = make_response(content.read())
    resp.headers["Content-Disposition"] = f"attachment; filename={fileName}.xlsx"
    resp.headers['Content-Type'] = 'application/x-xlsx'
    
    return resp


def auto_fit_columns(ws):
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter  # Get the column letter (e.g., A, B, C)
        for cell in column_cells:
            try:
                if cell.value:  # Skip empty cells
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2  # Add some padding
        ws.column_dimensions[column_letter].width = adjusted_width
        

class ExcelO:
    # ExcelO
    @staticmethod    
    def get_all_targets(targets):
        fileName = f"Vsi_naslovniki"
        
        # Create a new Workbook
        wb = Workbook()
        
        # Access the active sheet
        ws = wb.active
        
        # Rename the sheet (optional)
        ws.title = "Naslovniki"
        
        bold_font = Font(bold=True)      
        
        ws[f'A1'] = "Ime"
        ws['A1'].font = bold_font  
        ws[f'B1'] = "Email"
        ws['B1'].font = bold_font  
        ws[f'C1'] = "Aktiven ?"
        ws['C1'].font = bold_font  
        
        
        #table
        first_row = 2
        for target in targets:
            ws[f'A{first_row}'] = target['name']
            ws[f'B{first_row}'] = target['email']
            active = target['active']
            if active == True:
                active = "Aktiven"
            else:
                active = "Neaktiven"
            ws[f'C{first_row}'] = active
            
            first_row += 1
        
        # Automatically adjust column widths
        auto_fit_columns(ws)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        wordBook = makeResponse(output, fileName)
        
        return wordBook